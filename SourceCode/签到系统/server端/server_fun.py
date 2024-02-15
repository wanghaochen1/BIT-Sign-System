import sqlite3
from db import *
from openpyxl import load_workbook
import datetime
def check (con,stu_num,stu_password):
    #打开db.xlsx文件，并在sheet1中的第一列中寻找名字
    wb = load_workbook('db.xlsx')
    ws = wb['Log_in']
    for i in range(1,ws.max_row+1):
        if str(ws.cell(row=i,column=1).value) == str(stu_num) and str(ws.cell(row=i,column=2).value) == str(stu_password):
            last_col=len(ws[i])
            ws.cell(row=i,column=last_col+1).value = "✓"
            wb.save('db.xlsx')
            return True
        ret=select_table(con,"student",student_id=stu_num,student_pwd=stu_password)
    if ret:
        print("学生存在数据库")
        return True
    else:
        print("学生不存在数据库")
        return False   
    #数据库部分,目前还没有打开

   

def write_sign_in(con,stu_num, signtime):
    wb = load_workbook('db.xlsx')
    ws = wb[f'{stu_num}']
    new_row = [stu_num,signtime]
    ws.append(new_row)
    wb.save('db.xlsx')

    #数据库部分，目前可以选择使用
    ret=select_table(con,"student",student_id=stu_num)
    if ret:
        insert_table_qd(con,"qd",student_id=stu_num,sign_time=signtime,sign_out_time=-1.0)
        print("签到成功")
        return True
    else:
        print("学生不存在")
        return False

def allow_change_face(con,stu_num):
    wb = load_workbook('db.xlsx')
    ws = wb['Log_in']
    for i in range(1,ws.max_row+1):
        if str(ws.cell(row=i,column=1).value) == str(stu_num):
            return True
    ret=select_table(con,"student",student_id=stu_num)
    if ret:
        print("学生存在")
        return True
    else:
        print("学生不存在")
        return False   
    #数据库部分，目前可以选择使用

    
def write_sign_out(con,stu_num, sign_out_time):
    wb = load_workbook('db.xlsx')
    ws = wb[f'{stu_num}']
    for i in range(1,ws.max_row+1):
        if str(ws.cell(row=i,column=1).value) == str(stu_num)and not any(char.isdigit() for char in str(ws.cell(row=i,column=3).value)):
            ws.cell(row=i,column=3).value = sign_out_time
            wb.save('db.xlsx')
            return True
    print(f"\n学生{stu_num}没有签到,但是签退了,请联系他\n")        
    ret=select_table(con,"student",student_id=stu_num)
    if ret:
        update_table_qd(con,"qd",student_id=stu_num,index="sign_out_time",value=sign_out_time)
        print("签退成功")
        return True
    else:
        print("学生不存在")
        return False
    return False  
    
def user_image(con,stu_num):
    wb=load_workbook('db.xlsx')
    ws=wb[f'{stu_num}']
    data=[]
    for i in range(2, ws.max_row + 1):
        # 获取第2列和第3列的单元格的值
        col1_value = ws.cell(row=i, column=2).value
        col2_value = ws.cell(row=i, column=3).value

        # 如果值是 datetime 对象，将其转换为字符串
        if isinstance(col1_value, datetime.datetime):
            col1_value = col1_value.isoformat()
        if isinstance(col2_value, datetime.datetime):
            col2_value = col2_value.isoformat()

        # 将这两个值添加到一个列表中，然后添加到二维数组中
        data.append([col1_value, col2_value])
    
    ret=select_table(con,"student",student_id=stu_num)
    return data